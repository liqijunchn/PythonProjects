#%%
print("hello world")
#%%
from openpyxl import Workbook,load_workbook
wb=load_workbook("test.xlsx")
print(wb.get_sheet_names())

#%%
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active
ws1.title = "range names"
for row in range(1, 40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save(filename = dest_filename)
#%%
'''
something wrong with iter_rows
'''
from openpyxl import Workbook,load_workbook
wb=Workbook()
ws=wb.active
print(ws)
# for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
#     for cell in row:
#         print(cell)

#%%
from openpyxl import Workbook,load_workbook
wb=Workbook()
ws=wb.active
ws['C9'] = 'hello world'
tuple(ws.rows)
#save file
wb.save("test.xlsx")
#%%
'''
为什么empty_book.xlsx的data sheet 表单中没有值。
'''

from openpyxl import Workbook
# from openpyxl.compat import range
from openpyxl.utils import get_column_letter
wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active
ws1.title = "range names"
for row in range(1, 40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save(filename = dest_filename)
#%%
'''
read an existing workbook
'''
from openpyxl import load_workbook
wb = load_workbook(filename = 'empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)

#%%
'''
use number format
'''
import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# set date using a Python datetime
ws['A1'] = datetime.datetime(2018, 3, 12)

ws['A1'].number_format
 # You>>> # You can enable type inference on a case-by-case basis
wb.guess_types = True
# set percentage using a string followed by the percent sign
ws['B1'] = '3.14%'
wb.guess_types = False
ws['B1'].value
ws['B1'].number_format


import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# set date using a Python datetime
ws['A1'] = datetime.datetime(2010, 7, 21)